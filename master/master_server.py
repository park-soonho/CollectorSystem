"""
master/master_server.py

Master 서버
- 여러 Agent 에서 리소스를 수집해 Excel 로 저장
- Agent 의 API 포트(7000 or 7001)를 자동 탐지
- 웹 대시보드
    · HTTP  : 0.0.0.0:8001
    · HTTPS : 0.0.0.0:8443
- 수집 API 포트: 7000 (사용 중이면 자동으로 7001)
"""

import os
import sys
import ssl
import logging
import threading
import time
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional

import requests
import pandas as pd
import schedule
from flask import Flask, jsonify, render_template_string
from werkzeug.serving import make_server

# ── 경로 설정 ──────────────────────────────────────────
BASE_DIR   = Path(__file__).resolve().parent.parent
CERT_DIR   = BASE_DIR / "certs"
REPORT_DIR = BASE_DIR / "reports"
sys.path.insert(0, str(BASE_DIR))

from common.port_utils import resolve_api_port

# ── 환경 변수 ───────────────────────────────────────────
HTTP_PORT  = int(os.getenv("MASTER_HTTP_PORT",  "8001"))
HTTPS_PORT = int(os.getenv("MASTER_HTTPS_PORT", "8443"))
SSL_CERT   = os.getenv("SSL_CERT", str(CERT_DIR / "server.crt"))
SSL_KEY    = os.getenv("SSL_KEY",  str(CERT_DIR / "server.key"))

AGENT_TIMEOUT   = int(os.getenv("AGENT_TIMEOUT",   "10"))
COLLECT_EVERY_M = int(os.getenv("COLLECT_EVERY_M", "5"))   # 수집 주기(분)

# ── 로깅 ────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s - %(message)s",
)
logger = logging.getLogger("master")

# ── Master 자신의 API 포트 결정 ──────────────────────────
MASTER_API_PORT = resolve_api_port(primary=7000, fallback=7001)

REPORT_DIR.mkdir(parents=True, exist_ok=True)

# ── Flask 앱 ─────────────────────────────────────────────
app = Flask(__name__)

# 최신 수집 결과 메모리 캐시
_latest_resources: List[Dict] = []


# ────────────────────────────────────────────────────────
# Agent 포트 자동 탐지
# ────────────────────────────────────────────────────────
def detect_agent_port(base_url: str, candidates=(7000, 7001)) -> Optional[int]:
    """
    Agent 의 /api/port 엔드포인트를 통해 실제 API 포트를 확인.
    응답이 없으면 candidates 를 순서대로 직접 시도.
    """
    for port in candidates:
        try:
            r = requests.get(f"{base_url}:{port}/api/port", timeout=3)
            if r.status_code == 200:
                api_port = r.json().get("api_port", port)
                logger.info(f"  └─ Agent 포트 감지: {base_url}:{api_port}")
                return api_port
        except Exception:
            continue
    logger.warning(f"  └─ Agent 포트 감지 실패: {base_url}")
    return None


# ────────────────────────────────────────────────────────
# ResourceCollector
# ────────────────────────────────────────────────────────
class ResourceCollector:
    """Agent 서버들로부터 리소스를 수집하고 Excel 로 저장"""

    def __init__(self, agent_list: List[Dict[str, str]]):
        """
        agent_list 예시:
            [
              {"name": "web-01", "host": "http://192.168.1.10"},
              {"name": "db-01",  "host": "http://192.168.1.20"},
            ]
        host 에 포트를 포함하지 말 것. 포트는 자동 탐지.
        """
        self.agents  = agent_list
        self.timeout = AGENT_TIMEOUT

    # ── 단일 Agent 수집 ──────────────────────────────────
    def _fetch_one(self, agent: Dict[str, str]) -> Optional[Dict]:
        name = agent.get("name", "unknown")
        host = agent.get("host", "")

        port = detect_agent_port(host)
        if port is None:
            logger.error(f"✗ [{name}] 포트 탐지 실패, 건너뜀")
            return None

        url = f"{host}:{port}/api/resource"
        try:
            r = requests.get(url, timeout=self.timeout)
            if r.status_code == 200:
                logger.info(f"✓ [{name}] 수집 성공 (포트 {port})")
                return r.json()
            else:
                logger.error(f"✗ [{name}] HTTP {r.status_code}")
        except requests.exceptions.Timeout:
            logger.error(f"✗ [{name}] 타임아웃")
        except requests.exceptions.ConnectionError:
            logger.error(f"✗ [{name}] 연결 실패")
        except Exception as e:
            logger.error(f"✗ [{name}] 오류: {e}")
        return None

    # ── 전체 수집 ────────────────────────────────────────
    def collect_all(self) -> List[Dict]:
        logger.info("─" * 60)
        logger.info(f"리소스 수집 시작 ({len(self.agents)} 서버)")
        results = [self._fetch_one(a) for a in self.agents]
        ok = [r for r in results if r]
        logger.info(f"수집 완료: {len(ok)}/{len(self.agents)} 성공")
        return ok

    # ── DataFrame 변환 ───────────────────────────────────
    @staticmethod
    def to_dataframe(resources: List[Dict]) -> pd.DataFrame:
        rows = []
        for r in resources:
            rows.append({
                "서버명":         r.get("server_name", "N/A"),
                "API 포트":       r.get("api_port",    "N/A"),
                "수집시간":       r.get("timestamp",   "N/A"),
                "CPU 사용률(%)":  r.get("cpu",    {}).get("usage_percent", 0),
                "CPU 코어수":     r.get("cpu",    {}).get("core_count",    0),
                "메모리 전체(GB)": r.get("memory", {}).get("total_gb",      0),
                "메모리 사용(GB)": r.get("memory", {}).get("used_gb",       0),
                "메모리 사용률(%)": r.get("memory",{}).get("usage_percent", 0),
                "디스크 전체(GB)": r.get("disk",   {}).get("total_gb",      0),
                "디스크 사용(GB)": r.get("disk",   {}).get("used_gb",       0),
                "디스크 사용률(%)": r.get("disk",  {}).get("usage_percent", 0),
                "네트워크 송신(MB)": r.get("network",{}).get("bytes_sent_mb",0),
                "네트워크 수신(MB)": r.get("network",{}).get("bytes_recv_mb",0),
            })
        return pd.DataFrame(rows)

    # ── Excel 저장 ───────────────────────────────────────
    @staticmethod
    def save_excel(df: pd.DataFrame) -> Optional[Path]:
        ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = REPORT_DIR / f"server_resources_{ts}.xlsx"
        try:
            with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="리소스 현황", index=False)
                ws = writer.sheets["리소스 현황"]

                # 헤더 스타일
                from openpyxl.styles import PatternFill, Font, Alignment
                header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
                header_font = Font(color="FFFFFF", bold=True)
                for cell in ws[1]:
                    cell.fill      = header_fill
                    cell.font      = header_font
                    cell.alignment = Alignment(horizontal="center")

                # 열 너비 자동 조정
                for col in ws.columns:
                    max_len = max((len(str(c.value or "")) for c in col), default=10)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

                # 임계값 초과 행 강조 (CPU > 80%)
                from openpyxl.styles import fills
                warn_fill = PatternFill(fill_type="solid", fgColor="FFCCCC")
                cpu_col   = 4  # "CPU 사용률(%)" 열 번호 (1-based)
                for row in ws.iter_rows(min_row=2):
                    try:
                        if float(row[cpu_col - 1].value or 0) > 80:
                            for cell in row:
                                cell.fill = warn_fill
                    except (ValueError, TypeError):
                        pass

            logger.info(f"✓ Excel 저장: {filepath}")
            return filepath
        except Exception as e:
            logger.error(f"✗ Excel 저장 실패: {e}")
            return None

    # ── 통합 실행 ────────────────────────────────────────
    def run_once(self) -> Optional[Path]:
        global _latest_resources
        resources = self.collect_all()
        if not resources:
            logger.warning("수집된 데이터 없음")
            return None
        _latest_resources = resources
        df = self.to_dataframe(resources)
        return self.save_excel(df)


# ────────────────────────────────────────────────────────
# 웹 대시보드 (Flask) – 8001 / 8443
# ────────────────────────────────────────────────────────
DASHBOARD_HTML = """
<!DOCTYPE html>
<html lang="ko">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>CollectorSystem – 리소스 현황</title>
  <style>
    * { box-sizing: border-box; margin: 0; padding: 0; }
    body { font-family: 'Segoe UI', sans-serif; background: #0f1117; color: #e2e8f0; }
    header { background: #1a1f2e; padding: 20px 40px;
             border-bottom: 2px solid #2d3748; }
    header h1 { font-size: 1.5rem; color: #63b3ed; }
    header span { font-size: .85rem; color: #718096; }
    .grid { display: grid;
            grid-template-columns: repeat(auto-fill, minmax(320px, 1fr));
            gap: 20px; padding: 30px 40px; }
    .card { background: #1a1f2e; border-radius: 12px;
            padding: 20px; border: 1px solid #2d3748; }
    .card h2 { font-size: 1rem; color: #90cdf4; margin-bottom: 12px; }
    .card .ts { font-size: .75rem; color: #718096; margin-bottom: 14px; }
    .metric { display: flex; justify-content: space-between;
              align-items: center; padding: 6px 0;
              border-bottom: 1px solid #2d374830; }
    .metric:last-child { border-bottom: none; }
    .metric .label { font-size: .8rem; color: #a0aec0; }
    .metric .value { font-size: .9rem; font-weight: 600; }
    .bar-wrap { height: 6px; background: #2d3748;
                border-radius: 4px; width: 80px; }
    .bar { height: 6px; border-radius: 4px; background: #63b3ed; }
    .bar.warn  { background: #f6ad55; }
    .bar.alert { background: #fc8181; }
    .api-port  { font-size: .7rem; color: #68d391;
                 background: #1c3a2e; padding: 2px 8px;
                 border-radius: 9999px; display: inline-block; }
    .no-data { padding: 60px; text-align: center; color: #718096; }
  </style>
</head>
<body>
  <header>
    <h1>🖥️ CollectorSystem – 서버 리소스 현황</h1>
    <span>최종 수집: {{ collected_at }}</span>
  </header>
  <div class="grid">
    {% if servers %}
      {% for s in servers %}
      <div class="card">
        <h2>{{ s.server_name }}</h2>
        <span class="api-port">API :{{ s.api_port }}</span>
        <div class="ts">{{ s.timestamp }}</div>

        {% set cpu = s.cpu.usage_percent %}
        <div class="metric">
          <span class="label">CPU 사용률</span>
          <div style="display:flex;align-items:center;gap:8px">
            <span class="value">{{ cpu }}%</span>
            <div class="bar-wrap"><div class="bar {% if cpu>80 %}alert{% elif cpu>60 %}warn{% endif %}"
                 style="width:{{ cpu }}%"></div></div>
          </div>
        </div>

        {% set mem = s.memory.usage_percent %}
        <div class="metric">
          <span class="label">메모리 사용률</span>
          <div style="display:flex;align-items:center;gap:8px">
            <span class="value">{{ s.memory.used_gb }}GB / {{ s.memory.total_gb }}GB</span>
            <div class="bar-wrap"><div class="bar {% if mem>80 %}alert{% elif mem>60 %}warn{% endif %}"
                 style="width:{{ mem }}%"></div></div>
          </div>
        </div>

        {% set dsk = s.disk.usage_percent %}
        <div class="metric">
          <span class="label">디스크 사용률</span>
          <div style="display:flex;align-items:center;gap:8px">
            <span class="value">{{ s.disk.used_gb }}GB / {{ s.disk.total_gb }}GB</span>
            <div class="bar-wrap"><div class="bar {% if dsk>80 %}alert{% elif dsk>60 %}warn{% endif %}"
                 style="width:{{ dsk }}%"></div></div>
          </div>
        </div>

        <div class="metric">
          <span class="label">네트워크 송/수신</span>
          <span class="value">↑{{ s.network.bytes_sent_mb }}MB  ↓{{ s.network.bytes_recv_mb }}MB</span>
        </div>
      </div>
      {% endfor %}
    {% else %}
      <div class="no-data">수집된 데이터가 없습니다. 잠시 후 새로고침하세요.</div>
    {% endif %}
  </div>
</body>
</html>
"""


@app.route("/")
def dashboard():
    return render_template_string(
        DASHBOARD_HTML,
        servers=_latest_resources,
        collected_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    )


@app.route("/api/status", methods=["GET"])
def api_status():
    return jsonify({
        "master_api_port": MASTER_API_PORT,
        "http_port":       HTTP_PORT,
        "https_port":      HTTPS_PORT,
        "server_count":    len(_latest_resources),
        "last_collected":  _latest_resources[0].get("timestamp") if _latest_resources else None,
    })


@app.route("/api/resources", methods=["GET"])
def api_resources():
    return jsonify(_latest_resources)


# ────────────────────────────────────────────────────────
# 서버 스레드
# ────────────────────────────────────────────────────────
def _run_http():
    logger.info(f"HTTP  서버 시작: 0.0.0.0:{HTTP_PORT}")
    make_server("0.0.0.0", HTTP_PORT, app).serve_forever()


def _run_https():
    if not (Path(SSL_CERT).exists() and Path(SSL_KEY).exists()):
        logger.warning("SSL 인증서 없음 → HTTPS 비활성. certs/gen_certs.py 실행 필요")
        return
    ctx = ssl.SSLContext(ssl.PROTOCOL_TLS_SERVER)
    ctx.load_cert_chain(SSL_CERT, SSL_KEY)
    logger.info(f"HTTPS 서버 시작: 0.0.0.0:{HTTPS_PORT}")
    make_server("0.0.0.0", HTTPS_PORT, app, ssl_context=ctx).serve_forever()


# ────────────────────────────────────────────────────────
# Agent 서버 목록 (실제 환경에 맞게 수정)
# ────────────────────────────────────────────────────────
AGENT_SERVERS = [
    {"name": "web-server-1",  "host": "http://192.168.1.10"},
    {"name": "web-server-2",  "host": "http://192.168.1.11"},
    {"name": "db-server-1",   "host": "http://192.168.1.20"},
    {"name": "app-server-1",  "host": "http://192.168.1.30"},
]


# ────────────────────────────────────────────────────────
# 메인
# ────────────────────────────────────────────────────────
def main():
    collector = ResourceCollector(AGENT_SERVERS)

    logger.info("=" * 60)
    logger.info("  Master Server (CollectorSystem)")
    logger.info(f"  HTTP  대시보드: http://0.0.0.0:{HTTP_PORT}")
    logger.info(f"  HTTPS 대시보드: https://0.0.0.0:{HTTPS_PORT}")
    logger.info(f"  Master API    : 0.0.0.0:{MASTER_API_PORT}")
    logger.info(f"  Agent 수       : {len(AGENT_SERVERS)}")
    logger.info(f"  수집 주기      : {COLLECT_EVERY_M}분")
    logger.info("=" * 60)

    # 웹 서버 스레드 시작
    for fn in (_run_http, _run_https):
        threading.Thread(target=fn, daemon=True).start()

    # 초기 즉시 수집
    collector.run_once()

    # 주기적 스케줄
    schedule.every(COLLECT_EVERY_M).minutes.do(collector.run_once)

    logger.info(f"스케줄러 시작 (매 {COLLECT_EVERY_M}분 수집)")
    try:
        while True:
            schedule.run_pending()
            time.sleep(1)
    except KeyboardInterrupt:
        logger.info("Master 서버 종료")


if __name__ == "__main__":
    main()
